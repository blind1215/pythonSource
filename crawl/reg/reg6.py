import re
import openpyxl


# train.xlsx 파일을 읽어서 4개의 시티를 만들어 각 행의 정보 복사하기
# Mr(sheet 명은 남성), Miss(sheet 명은 미혼여성)
# Mrs(sheet 명은 기혼여성), Others(sheet 명은 기타)

# 엑셀파일 읽기
work_book = openpyxl.load_workbook("./crawl/data/train.xlsx")
# 시트 활성화
sheet1 = work_book.active


# 성별 정보가 추가된 엑셀 파일 작성
wb = openpyxl.Workbook()
# 시트 4(기본시트 + 추가3)개를 추가
sheet_man = wb.active
sheet_man.column_dimensions["D"].width = 70
sheet_man.title = "남성"

sheet_solo_women = wb.create_sheet(title="미혼여성")
sheet_solo_women.column_dimensions["D"].width = 70

sheet_married_women = wb.create_sheet(title="기혼여성")
sheet_married_women.column_dimensions["D"].width = 70

sheet_others = wb.create_sheet(title="기타")
sheet_others.column_dimensions["D"].width = 70

pattern = re.compile(r" [A-Za-z]+\.")

# 생존자수와 사망자 수
man_survived, man_unsurvived = 0, 0
single_survived, single_unsurvived = 0, 0
married_survived, married_unsurvived = 0, 0
others_survived, others_unsurvived = 0, 0


# 필드명 옮기는 작업함
for each_row in sheet1.rows:
    # 이름을 읽어 패턴과 일치하는 데이터 담기
    data = pattern.findall(each_row[3].value)

    if each_row[0].row == 1:
        titles = []
        for each_item in each_row:
            titles.append(each_item.value)
        sheet_man.append(titles)
        sheet_solo_women.append(titles)
        sheet_married_women.append(titles)
        sheet_others.append(titles)
    else:
        # 2번행부터 이름 열 데이터를 읽은 후  각 성별에 맞춰 시트에 붙여넣기
        if len(data) > 0:
            if data[0] == " Mr.":
                sheet_man.append([each_item.value for each_item in each_row])
                if each_row[1].value == 1:  # 생존자
                    man_survived += 1
                else:
                    man_unsurvived += 1
            elif data[0] == " Miss.":
                sheet_solo_women.append([each_item.value for each_item in each_row])
                if each_row[1].value == 1:  # 생존자
                    single_survived += 1
                else:
                    single_unsurvived += 1
            elif data[0] == " Mrs.":
                sheet_married_women.append([each_item.value for each_item in each_row])
                if each_row[1].value == 1:  # 생존자
                    married_survived += 1
                else:
                    married_unsurvived += 1
            else:
                sheet_others.append([each_item.value for each_item in each_row])
                if each_row[1].value == 1:  # 생존자
                    others_survived += 1
                else:
                    others_unsurvived += 1


# 5번째 시트 생성 후 보고서 기록
sheet_report = wb.create_sheet(title="보고서")

# 필드명 추가
sheet_report.append(["분류", "생존자수", "사망자수", "생존률"])

# 남성 생존율 계산
survived_rate = man_survived / (man_survived + man_unsurvived) * 100
sheet_report.append(["남성", man_survived, man_unsurvived, survived_rate])

# 미혼여성 생존율 계산
survived_rate = single_survived / (single_survived + single_unsurvived) * 100
sheet_report.append(["미혼여성", single_survived, single_unsurvived, survived_rate])

# 기혼여성 생존율 계산
survived_rate = married_survived / (married_survived + married_unsurvived) * 100
sheet_report.append(["기혼여성", married_survived, married_unsurvived, survived_rate])

# 기타 생존율 계산
survived_rate = others_survived / (others_survived + others_unsurvived) * 100
sheet_report.append(["기타", others_survived, others_unsurvived, survived_rate])


wb.save("./crawl/data/train_gender.xlsx")
wb.close()
work_book.close()
